import logging
from pathlib import Path

import openpyxl
from django.conf import settings
from django.db.models import Count
from robots.models import Robot
from utils.date import get_last_monday

logger = logging.getLogger("django")


def delete_report(report_filepath="excel_reports/summary_for_week.xlsx"):
    logger.info("delete_report CALLED")
    report_file = Path(settings.BASE_DIR) / report_filepath
    report_file.unlink(missing_ok=True)


def create_report(report_filepath="excel_reports/summary_for_week.xlsx"):
    logger.info("create_report CALLED")
    WeekReport(report_filepath=report_filepath).create()


class WeekReport:
    def __init__(self, report_filepath):
        self.report_filepath = Path(settings.BASE_DIR) / report_filepath
        self.wb = openpyxl.Workbook()

    def create(self):
        """create summary file for recently created robots"""

        last_monday = get_last_monday()
        del self.wb["Sheet"]

        models_count = (
            Robot.objects.filter(created__gt=last_monday)
            .values("model")
            .annotate(total=Count("model"))
            .order_by("total")
        )
        for model_data in models_count:
            ws = self.wb.create_sheet(model_data["model"])

            ws["A1"] = "Модель"
            ws["B1"] = "Версия"
            ws["C1"] = "Количество за неделю"

            versions_count = (
                Robot.objects.filter(created__gt=last_monday, model=model_data["model"])
                .values("version")
                .annotate(total=Count("version"))
                .order_by("total")
            )
            for n, version_data in enumerate(versions_count, start=1):
                ws[f"A{n + 1}"] = model_data["model"]
                ws[f"B{n + 1}"] = version_data["version"]
                ws[f"C{n + 1}"] = version_data["total"]

        if models_count:
            self.wb.save(self.report_filepath)
            logger.info(f"{self.report_filepath} created. Status: OK")
        else:
            logger.info(f"{self.report_filepath} not created. No robots found")
